"""Excel帳票出力モジュール（7シート構成）"""
import io
from datetime import date, datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference


# ─── スタイル定数 ────────────────────────────────────────────

HEADER_BG = "1F4E79"    # 濃い青
SUB_HEADER_BG = "2E75B6"  # 中青
ACCENT_BG = "DEEAF1"    # 薄い青（交互行）
PAID_BG = "E2EFDA"      # 薄い緑（支払済み）
OVERDUE_BG = "FCE4D6"   # 薄い赤（期限超過）
DUE_SOON_BG = "FFF2CC"  # 薄い黄（期限間近）
TOTAL_BG = "F2F2F2"     # グレー（合計行）

H_FONT = Font(name="游ゴシック", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="游ゴシック", bold=True, size=14)
NORMAL_FONT = Font(name="游ゴシック", size=10)
BOLD_FONT = Font(name="游ゴシック", bold=True, size=10)
TOTAL_FONT = Font(name="游ゴシック", bold=True, size=10, color="1F4E79")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

AMOUNT_FMT = '#,##0'
DATE_FMT = 'YYYY/MM/DD'

THIN = Side(style="thin", color="BFBFBF")
MEDIUM = Side(style="medium", color="1F4E79")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BORDER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _set_row_style(ws, row, fill=None, font=None, border=True, height=18):
    for cell in row:
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        else:
            cell.font = NORMAL_FONT
        if border:
            cell.border = THIN_BORDER
        cell.alignment = LEFT
    if height:
        ws.row_dimensions[row[0].row].height = height


def _auto_col_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                # 日本語文字は2倍幅として計算
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 2, max_w))


def _write_title(ws, title: str, subtitle: str = "", row: int = 1):
    ws.merge_cells(f"A{row}:K{row}")
    c = ws.cell(row=row, column=1, value=title)
    c.font = TITLE_FONT
    c.alignment = CENTER
    c.fill = _make_fill(HEADER_BG)
    c.font = Font(name="游ゴシック", bold=True, size=14, color="FFFFFF")
    ws.row_dimensions[row].height = 28
    if subtitle:
        ws.merge_cells(f"A{row+1}:K{row+1}")
        c2 = ws.cell(row=row + 1, column=1, value=subtitle)
        c2.font = Font(name="游ゴシック", size=10, color="595959")
        c2.alignment = CENTER
        ws.row_dimensions[row + 1].height = 16


def _get_status_fill(row_data: dict) -> PatternFill | None:
    status = row_data.get("status", "")
    adj_date = row_data.get("adjusted_date", "")
    today = date.today().isoformat()
    if status == "paid":
        return _make_fill(PAID_BG)
    if status == "unpaid" and adj_date and adj_date < today:
        return _make_fill(OVERDUE_BG)
    from datetime import timedelta
    soon = (date.today() + timedelta(days=3)).isoformat()
    if status == "unpaid" and adj_date and today <= adj_date <= soon:
        return _make_fill(DUE_SOON_BG)
    return None


STATUS_LABELS = {"paid": "支払済み", "unpaid": "未払い"}
TYPE_LABELS = {"fixed": "固定", "variable": "変動"}


# ─── シート1: 支払い一覧 ───────────────────────────────────────

def _sheet_payment_list(wb: Workbook, df: pd.DataFrame, year: int, month: int):
    ws = wb.create_sheet("支払い一覧")
    ws.freeze_panes = "A5"

    _write_title(ws, f"支払い一覧　{year}年{month}月", f"出力日: {date.today().strftime('%Y/%m/%d')}")

    headers = ["No.", "支払先", "支払内容", "固定/変動", "支払日", "金額",
               "支払方法", "ステータス", "カテゴリ", "備考", "調整後支払日"]
    col_widths = [5, 18, 20, 9, 7, 14, 12, 10, 14, 20, 12]

    header_row = 4
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.fill = _make_fill(SUB_HEADER_BG)
        c.font = H_FONT
        c.alignment = CENTER
        c.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[header_row].height = 22

    total_amount = 0
    paid_amount = 0
    for ri, (_, row) in enumerate(df.iterrows(), 1):
        r = header_row + ri
        status_label = STATUS_LABELS.get(row.get("status", ""), row.get("status", ""))
        type_label = TYPE_LABELS.get(row.get("payment_type", ""), row.get("payment_type", ""))
        adj = row.get("adjusted_date", "")
        try:
            adj_fmt = datetime.strptime(adj, "%Y-%m-%d").strftime("%Y/%m/%d") if adj else ""
        except Exception:
            adj_fmt = adj

        values = [
            ri,
            row.get("payee", ""),
            row.get("description", ""),
            type_label,
            row.get("payment_day", ""),
            row.get("amount", 0),
            row.get("payment_method", ""),
            status_label,
            row.get("category", ""),
            row.get("notes", ""),
            adj_fmt,
        ]
        row_data = dict(row)
        fill = _get_status_fill(row_data)
        if not fill:
            fill = _make_fill(ACCENT_BG) if ri % 2 == 0 else None

        for ci, val in enumerate(values, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = NORMAL_FONT
            c.border = THIN_BORDER
            c.alignment = CENTER if ci in (1, 4, 5, 7, 8) else LEFT
            if fill:
                c.fill = fill
            if ci == 6:
                c.number_format = AMOUNT_FMT
                c.alignment = RIGHT

        amount = row.get("amount", 0) or 0
        total_amount += amount
        if row.get("status") == "paid":
            paid_amount += amount

    total_row = header_row + len(df) + 1
    ws.merge_cells(f"A{total_row}:E{total_row}")
    c = ws.cell(row=total_row, column=1, value="合　計")
    c.font = TOTAL_FONT
    c.fill = _make_fill(TOTAL_BG)
    c.alignment = CENTER
    c.border = HEADER_BORDER

    tc = ws.cell(row=total_row, column=6, value=total_amount)
    tc.font = TOTAL_FONT
    tc.fill = _make_fill(TOTAL_BG)
    tc.number_format = AMOUNT_FMT
    tc.alignment = RIGHT
    tc.border = HEADER_BORDER

    paid_row = total_row + 1
    ws.merge_cells(f"A{paid_row}:E{paid_row}")
    c2 = ws.cell(row=paid_row, column=1, value="支払済み合計")
    c2.font = Font(name="游ゴシック", bold=True, size=10, color="375623")
    c2.fill = _make_fill(PAID_BG)
    c2.alignment = CENTER
    c2.border = HEADER_BORDER

    tc2 = ws.cell(row=paid_row, column=6, value=paid_amount)
    tc2.font = Font(name="游ゴシック", bold=True, size=10, color="375623")
    tc2.fill = _make_fill(PAID_BG)
    tc2.number_format = AMOUNT_FMT
    tc2.alignment = RIGHT
    tc2.border = HEADER_BORDER

    unpaid_row = total_row + 2
    ws.merge_cells(f"A{unpaid_row}:E{unpaid_row}")
    c3 = ws.cell(row=unpaid_row, column=1, value="未払い合計")
    c3.font = Font(name="游ゴシック", bold=True, size=10, color="833C00")
    c3.fill = _make_fill(OVERDUE_BG)
    c3.alignment = CENTER
    c3.border = HEADER_BORDER

    tc3 = ws.cell(row=unpaid_row, column=6, value=total_amount - paid_amount)
    tc3.font = Font(name="游ゴシック", bold=True, size=10, color="833C00")
    tc3.fill = _make_fill(OVERDUE_BG)
    tc3.number_format = AMOUNT_FMT
    tc3.alignment = RIGHT
    tc3.border = HEADER_BORDER


# ─── シート2: 月別一覧 ────────────────────────────────────────

def _sheet_monthly_list(wb: Workbook, df_year: pd.DataFrame, year: int):
    ws = wb.create_sheet("月別一覧")
    ws.freeze_panes = "A5"

    _write_title(ws, f"月別支払い一覧　{year}年", f"出力日: {date.today().strftime('%Y/%m/%d')}")

    headers = ["月", "No.", "支払先", "支払内容", "固定/変動", "支払日",
               "金額", "支払方法", "ステータス", "カテゴリ", "備考"]
    col_widths = [6, 5, 18, 20, 9, 7, 14, 12, 10, 14, 20]

    header_row = 4
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.fill = _make_fill(SUB_HEADER_BG)
        c.font = H_FONT
        c.alignment = CENTER
        c.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[header_row].height = 22

    r = header_row
    for month in range(1, 13):
        month_df = df_year[df_year["month"] == month]
        if month_df.empty:
            continue
        month_start_row = r + 1
        for ri, (_, row) in enumerate(month_df.iterrows(), 1):
            r += 1
            type_label = TYPE_LABELS.get(row.get("payment_type", ""), "")
            status_label = STATUS_LABELS.get(row.get("status", ""), "")
            fill = _get_status_fill(dict(row))
            if not fill:
                fill = _make_fill(ACCENT_BG) if ri % 2 == 0 else None

            values = [
                f"{month}月" if ri == 1 else "",
                ri,
                row.get("payee", ""),
                row.get("description", ""),
                type_label,
                row.get("payment_day", ""),
                row.get("amount", 0),
                row.get("payment_method", ""),
                status_label,
                row.get("category", ""),
                row.get("notes", ""),
            ]
            for ci, val in enumerate(values, 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font = NORMAL_FONT
                c.border = THIN_BORDER
                c.alignment = CENTER if ci in (1, 2, 5, 6, 8, 9) else LEFT
                if fill:
                    c.fill = fill
                if ci == 7:
                    c.number_format = AMOUNT_FMT
                    c.alignment = RIGHT

        month_total = month_df["amount"].sum()
        r += 1
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(row=r, column=1, value=f"{month}月 合計")
        c.font = TOTAL_FONT
        c.fill = _make_fill(TOTAL_BG)
        c.alignment = RIGHT
        c.border = HEADER_BORDER
        tc = ws.cell(row=r, column=7, value=month_total)
        tc.font = TOTAL_FONT
        tc.fill = _make_fill(TOTAL_BG)
        tc.number_format = AMOUNT_FMT
        tc.alignment = RIGHT
        tc.border = HEADER_BORDER
        r += 1  # 空行


# ─── シート3: 支払先一覧 ─────────────────────────────────────

def _sheet_payee_list(wb: Workbook, df: pd.DataFrame, year: int):
    ws = wb.create_sheet("支払先一覧")

    _write_title(ws, f"支払先別集計　{year}年", f"出力日: {date.today().strftime('%Y/%m/%d')}")

    headers = ["支払先", "件数", "支払総額", "支払済み額", "未払い額", "固定件数", "変動件数"]
    col_widths = [20, 8, 16, 16, 16, 10, 10]

    header_row = 4
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.fill = _make_fill(SUB_HEADER_BG)
        c.font = H_FONT
        c.alignment = CENTER
        c.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[header_row].height = 22

    payee_group = df.groupby("payee").agg(
        count=("id", "count"),
        total=("amount", "sum"),
        paid=("amount", lambda x: x[df.loc[x.index, "status"] == "paid"].sum()),
        fixed=("payment_type", lambda x: (x == "fixed").sum()),
        variable=("payment_type", lambda x: (x == "variable").sum()),
    ).reset_index()

    for ri, (_, row) in enumerate(payee_group.iterrows(), 1):
        r = header_row + ri
        fill = _make_fill(ACCENT_BG) if ri % 2 == 0 else None
        values = [
            row["payee"],
            row["count"],
            row["total"],
            row["paid"],
            row["total"] - row["paid"],
            row["fixed"],
            row["variable"],
        ]
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = NORMAL_FONT
            c.border = THIN_BORDER
            c.alignment = RIGHT if ci > 1 else LEFT
            if fill:
                c.fill = fill
            if ci in (3, 4, 5):
                c.number_format = AMOUNT_FMT


# ─── シート4: 集計シート ─────────────────────────────────────

def _sheet_summary(wb: Workbook, df: pd.DataFrame, year: int, month: int):
    ws = wb.create_sheet("集計シート")
    _write_title(ws, f"支払い集計　{year}年{month}月", f"出力日: {date.today().strftime('%Y/%m/%d')}")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10

    def write_section(start_row: int, title: str, rows: list[tuple]) -> int:
        r = start_row
        ws.merge_cells(f"A{r}:C{r}")
        c = ws.cell(row=r, column=1, value=title)
        c.fill = _make_fill(SUB_HEADER_BG)
        c.font = H_FONT
        c.alignment = LEFT
        c.border = HEADER_BORDER
        ws.row_dimensions[r].height = 20
        r += 1
        for label, value, is_amount in rows:
            c1 = ws.cell(row=r, column=1, value=label)
            c1.font = NORMAL_FONT
            c1.border = THIN_BORDER
            c1.alignment = LEFT
            c2 = ws.cell(row=r, column=2, value=value)
            c2.font = BOLD_FONT
            c2.border = THIN_BORDER
            c2.alignment = RIGHT
            if is_amount:
                c2.number_format = AMOUNT_FMT
            c3 = ws.cell(row=r, column=3, value="円" if is_amount else "件")
            c3.font = NORMAL_FONT
            c3.border = THIN_BORDER
            c3.alignment = CENTER
            r += 1
        r += 1
        return r

    total = df["amount"].sum()
    paid_df = df[df["status"] == "paid"]
    unpaid_df = df[df["status"] == "unpaid"]
    today = date.today().isoformat()
    overdue_df = unpaid_df[unpaid_df["adjusted_date"] < today]

    r = 4
    r = write_section(r, "■ 全体集計", [
        ("支払い件数", len(df), False),
        ("支払い総額", total, True),
        ("支払済み件数", len(paid_df), False),
        ("支払済み額", paid_df["amount"].sum(), True),
        ("未払い件数", len(unpaid_df), False),
        ("未払い額", unpaid_df["amount"].sum(), True),
        ("期限超過件数", len(overdue_df), False),
        ("期限超過額", overdue_df["amount"].sum(), True),
    ])

    r = write_section(r, "■ 固定/変動別", [
        ("固定支払い件数", len(df[df["payment_type"] == "fixed"]), False),
        ("固定支払い総額", df[df["payment_type"] == "fixed"]["amount"].sum(), True),
        ("変動支払い件数", len(df[df["payment_type"] == "variable"]), False),
        ("変動支払い総額", df[df["payment_type"] == "variable"]["amount"].sum(), True),
    ])

    method_group = df.groupby("payment_method")["amount"].sum().reset_index()
    method_rows = [(row["payment_method"] or "未設定", row["amount"], True)
                   for _, row in method_group.iterrows()]
    r = write_section(r, "■ 支払方法別", method_rows)


# ─── シート5: 月別集計 ────────────────────────────────────────

def _sheet_monthly_summary(wb: Workbook, df_year: pd.DataFrame, year: int):
    ws = wb.create_sheet("月別集計")
    _write_title(ws, f"月別集計　{year}年", f"出力日: {date.today().strftime('%Y/%m/%d')}")

    headers = ["月", "件数", "支払総額", "支払済み額", "未払い額", "達成率(%)"]
    col_widths = [8, 8, 16, 16, 16, 12]

    header_row = 4
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.fill = _make_fill(SUB_HEADER_BG)
        c.font = H_FONT
        c.alignment = CENTER
        c.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[header_row].height = 22

    for month in range(1, 13):
        mdf = df_year[df_year["month"] == month]
        r = header_row + month
        fill = _make_fill(ACCENT_BG) if month % 2 == 0 else None
        total = mdf["amount"].sum() if not mdf.empty else 0
        paid = mdf[mdf["status"] == "paid"]["amount"].sum() if not mdf.empty else 0
        ratio = round(paid / total * 100, 1) if total > 0 else 0
        values = [f"{month}月", len(mdf), total, paid, total - paid, ratio]
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = NORMAL_FONT
            c.border = THIN_BORDER
            c.alignment = CENTER if ci in (1, 2, 6) else RIGHT
            if fill:
                c.fill = fill
            if ci in (3, 4, 5):
                c.number_format = AMOUNT_FMT

    total_row = header_row + 13
    ws.merge_cells(f"A{total_row}:B{total_row}")
    c = ws.cell(row=total_row, column=1, value="年間合計")
    c.font = TOTAL_FONT
    c.fill = _make_fill(TOTAL_BG)
    c.alignment = CENTER
    c.border = HEADER_BORDER
    for ci, col in [(3, "amount"), (4, "paid"), (5, "unpaid")]:
        if col == "paid":
            val = df_year[df_year["status"] == "paid"]["amount"].sum()
        elif col == "unpaid":
            val = df_year[df_year["status"] == "unpaid"]["amount"].sum()
        else:
            val = df_year["amount"].sum()
        tc = ws.cell(row=total_row, column=ci, value=val)
        tc.font = TOTAL_FONT
        tc.fill = _make_fill(TOTAL_BG)
        tc.number_format = AMOUNT_FMT
        tc.alignment = RIGHT
        tc.border = HEADER_BORDER


# ─── シート6: 年間集計 ────────────────────────────────────────

def _sheet_annual_summary(wb: Workbook, df_year: pd.DataFrame, year: int):
    ws = wb.create_sheet("年間集計")
    _write_title(ws, f"年間集計　{year}年", f"出力日: {date.today().strftime('%Y/%m/%d')}")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10

    def cell(r, c, v, bold=False, fill_hex=None, num_fmt=None, align=LEFT):
        cell_obj = ws.cell(row=r, column=c, value=v)
        cell_obj.font = BOLD_FONT if bold else NORMAL_FONT
        cell_obj.border = THIN_BORDER
        cell_obj.alignment = align
        if fill_hex:
            cell_obj.fill = _make_fill(fill_hex)
        if num_fmt:
            cell_obj.number_format = num_fmt
        return cell_obj

    r = 4
    ws.merge_cells(f"A{r}:C{r}")
    c = ws.cell(row=r, column=1, value="■ 年間サマリー")
    c.fill = _make_fill(SUB_HEADER_BG)
    c.font = H_FONT
    c.alignment = LEFT
    c.border = HEADER_BORDER
    ws.row_dimensions[r].height = 20
    r += 1

    total = df_year["amount"].sum()
    paid = df_year[df_year["status"] == "paid"]["amount"].sum()
    items = [
        ("年間支払い件数", len(df_year), False),
        ("年間支払い総額", total, True),
        ("支払済み額", paid, True),
        ("未払い額", total - paid, True),
        ("支払達成率", f"{round(paid/total*100, 1) if total > 0 else 0}%", False),
        ("月平均支払い額", round(total / 12, 0), True),
        ("固定支払い総額", df_year[df_year["payment_type"] == "fixed"]["amount"].sum(), True),
        ("変動支払い総額", df_year[df_year["payment_type"] == "variable"]["amount"].sum(), True),
    ]
    for label, val, is_amt in items:
        cell(r, 1, label)
        c2 = cell(r, 2, val, bold=True, align=RIGHT)
        if is_amt:
            c2.number_format = AMOUNT_FMT
        cell(r, 3, "円" if is_amt else ("件" if not isinstance(val, str) else ""), align=CENTER)
        r += 1


# ─── シート7: カテゴリ別集計 ─────────────────────────────────

def _sheet_category_summary(wb: Workbook, df_year: pd.DataFrame, year: int):
    ws = wb.create_sheet("カテゴリ別集計")
    _write_title(ws, f"カテゴリ別集計　{year}年", f"出力日: {date.today().strftime('%Y/%m/%d')}")

    headers = ["カテゴリ", "件数", "支払総額", "支払済み額", "未払い額", "構成比(%)"]
    col_widths = [20, 8, 16, 16, 16, 12]

    header_row = 4
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=header_row, column=ci, value=h)
        c.fill = _make_fill(SUB_HEADER_BG)
        c.font = H_FONT
        c.alignment = CENTER
        c.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[header_row].height = 22

    df_copy = df_year.copy()
    df_copy["category"] = df_copy["category"].fillna("未分類").replace("", "未分類")
    total_all = df_copy["amount"].sum()

    cat_group = df_copy.groupby("category").agg(
        count=("id", "count"),
        total=("amount", "sum"),
    ).reset_index().sort_values("total", ascending=False)

    paid_by_cat = df_copy[df_copy["status"] == "paid"].groupby("category")["amount"].sum()

    for ri, (_, row) in enumerate(cat_group.iterrows(), 1):
        r = header_row + ri
        fill = _make_fill(ACCENT_BG) if ri % 2 == 0 else None
        cat = row["category"]
        total_cat = row["total"]
        paid_cat = paid_by_cat.get(cat, 0)
        ratio = round(total_cat / total_all * 100, 1) if total_all > 0 else 0
        values = [cat, row["count"], total_cat, paid_cat, total_cat - paid_cat, ratio]
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = NORMAL_FONT
            c.border = THIN_BORDER
            c.alignment = LEFT if ci == 1 else (CENTER if ci in (2, 6) else RIGHT)
            if fill:
                c.fill = fill
            if ci in (3, 4, 5):
                c.number_format = AMOUNT_FMT

    total_r = header_row + len(cat_group) + 1
    ws.merge_cells(f"A{total_r}:B{total_r}")
    c = ws.cell(row=total_r, column=1, value="合　計")
    c.font = TOTAL_FONT
    c.fill = _make_fill(TOTAL_BG)
    c.alignment = RIGHT
    c.border = HEADER_BORDER
    tc = ws.cell(row=total_r, column=3, value=total_all)
    tc.font = TOTAL_FONT
    tc.fill = _make_fill(TOTAL_BG)
    tc.number_format = AMOUNT_FMT
    tc.alignment = RIGHT
    tc.border = HEADER_BORDER


# ─── メイン出力関数 ────────────────────────────────────────────

def export_to_excel(
    df: pd.DataFrame,
    df_year: pd.DataFrame,
    year: int,
    month: int,
) -> bytes:
    """
    7シート構成のExcelを生成してbytesで返す
    """
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    _sheet_payment_list(wb, df, year, month)
    _sheet_monthly_list(wb, df_year, year)
    _sheet_payee_list(wb, df_year, year)
    _sheet_summary(wb, df, year, month)
    _sheet_monthly_summary(wb, df_year, year)
    _sheet_annual_summary(wb, df_year, year)
    _sheet_category_summary(wb, df_year, year)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
